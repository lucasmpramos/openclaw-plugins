#!/usr/bin/env python3
"""Parse Arion Bank xlsx statement and output structured JSON for insertion.

Output format matches Google Sheet columns A-T exactly:
  A: Date, B: Amount ISK, C: Balance ISK, D: Amount USD, E: Balance USD,
  F: Description, G: Receipt No., H: Reference, I: Type (EN), J: Category,
  K-T: additional fields

Usage: parse_bank_statement.py <file.xlsx> [--rate 0.00807]
"""

import sys
import json
from datetime import datetime
from openpyxl import load_workbook


def safe(row, idx, default=''):
    """Safely get value from row, handling short rows."""
    val = row[idx] if len(row) > idx else default
    return val if val is not None else default


def parse(filepath, exchange_rate=None):
    wb = load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (contains 'Dagsetning')
    header_idx = None
    for i, row in enumerate(rows[:6]):
        if row and 'Dagsetning' in str(row):
            header_idx = i
            break

    if header_idx is None:
        print("ERROR: Could not find header row with 'Dagsetning'", file=sys.stderr)
        sys.exit(1)

    data_rows = rows[header_idx + 1:]

    records = []
    for row in data_rows:
        if not row[0]:
            continue

        date = row[0].strftime('%d.%m.%Y') if isinstance(row[0], datetime) else str(row[0])
        amount_isk = row[1] or 0
        balance_isk = row[2] or 0
        description = str(safe(row, 4, '')).strip()
        type_is = str(safe(row, 7, ''))

        # Build sheet row (A-T, 20 columns)
        sheet_row = [
            date,                                    # A: Date
            f"{amount_isk:,.2f}",                    # B: Amount ISK
            f"{balance_isk:,.2f}",                   # C: Balance ISK
            '',                                      # D: Amount USD (filled below)
            '',                                      # E: Balance USD (filled below)
            description,                              # F: Description
            str(safe(row, 5, '')),                    # G: Receipt No.
            str(safe(row, 6, '')),                    # H: Reference
            translate_type(type_is),                  # I: Type (translated to EN)
            '',                                      # J: Category (filled by categorize.py)
            str(safe(row, 8, '')),                    # K: Transaction Key
            str(safe(row, 9, '')),                    # L: Interest Day
            str(safe(row, 10, '')),                   # M: Text Key
            str(safe(row, 11, '')),                   # N: Payment Bank
            str(safe(row, 12, '')),                   # O: Batch No.
            str(safe(row, 13, '')),                   # P: Interest Date
            str(safe(row, 14, '')),                   # Q: Recipient/Payer Name
            str(safe(row, 15, '')),                   # R: Number
            str(safe(row, 16, '')),                   # S: Recipient/Payer ID
            str(safe(row, 17, '')),                   # T: Unique Key (dedup key)
        ]

        # USD conversion
        if exchange_rate:
            sheet_row[3] = f"{round(amount_isk * exchange_rate, 2):,.2f}"
            sheet_row[4] = f"{round(balance_isk * exchange_rate, 2):,.2f}"

        records.append(sheet_row)

    # Sort by date descending (newest first)
    records.sort(key=lambda r: r[0], reverse=True)

    return records


def translate_type(type_is):
    """Translate Icelandic transaction type to English."""
    mapping = {
        'Símgreiðsla': 'Wire Transfer',
        'Debitkortafærsla': 'Debit card',
        'Reikningur': 'Invoice',
        'Innheimtukrafa - kostnaður': 'Collection fee',
        'Vaxagreiðsla': 'Interest',
        'Þjónustugjald': 'Service fee',
        'Leiging': 'Leasing',
        'Krafanúmer': 'Collection',
        'Launagreiðsla': 'Payroll',
        'Innborgun': 'Deposit',
    }
    return mapping.get(type_is, type_is)


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: parse_bank_statement.py <file.xlsx> [--rate 0.00807]")
        sys.exit(1)

    filepath = sys.argv[1]
    rate = None

    if '--rate' in sys.argv:
        rate = float(sys.argv[sys.argv.index('--rate') + 1])

    records = parse(filepath, rate)

    if not records:
        print("No records found.")
        sys.exit(0)

    dates = [r[0] for r in records]
    print(f"Parsed {len(records)} transactions: {dates[-1]} to {dates[0]}")

    outpath = filepath.replace('.xlsx', '_parsed.json')
    if outpath == filepath:
        outpath = filepath + '_parsed.json'

    with open(outpath, 'w') as f:
        json.dump(records, f, indent=2)
    print(f"Saved to {outpath}")
